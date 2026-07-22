import os
import time
import warnings
import csv
from collections import defaultdict
import random
import time

import openpyxl
import requests
from dotenv import load_dotenv, set_key

def read_excel(filename):
    workbook = openpyxl.load_workbook(filename)

    addresses_sheet = workbook["Адреса"]
    coordinates_sheet = workbook["Координаты"]

    addresses = [
        row[0]
        for row in addresses_sheet.iter_rows(min_row=2, values_only=True)
        if row[0]
    ]

    coordinates = [
        row[0]
        for row in coordinates_sheet.iter_rows(min_row=2, values_only=True)
        if row[0]
    ]

    return addresses, coordinates


def save_result(service, operation, request, result, quality, status, elapsed):

    file_exists = os.path.exists("results.csv")

    with open("results.csv", "a", newline="", encoding="utf-8-sig") as file:

        writer = csv.writer(file)

        if not file_exists:
            writer.writerow([
                "Service",
                "Operation",
                "Request",
                "Result",
                "Quality",
                "HTTP Status",
                "Response Time (ms)"
            ])

        writer.writerow([
            service,
            operation,
            request,
            result,
            quality,
            status,
            round(elapsed * 1000, 2)
        ])

def update_gigachat_token():

    expire = int(os.getenv("GIGACHAT_TIME") or 0)

    if expire > int(time.time()):
        return

    headers = {
        "Authorization": f"Basic {os.getenv('GIGACHAT_AUTH_KEY')}",
        "RqUID": "12345678-1234-1234-1234-123456789012",
        "Content-Type": "application/x-www-form-urlencoded",
    }

    response = requests.post(
        "https://ngw.devices.sberbank.ru:9443/api/v2/oauth",
        headers=headers,
        data={
            "scope": os.getenv("GIGACHAT_SCOPE")
        },
        verify=False,
    )

    data = response.json()

    set_key(".env", "GIGACHAT_TOKEN", data["access_token"])
    set_key(".env", "GIGACHAT_TIME", str(data["expires_at"])[:-3])

    print("GigaChat token updated.")

def request_yandex(query, is_address):
    if is_address:
        query = query.replace(" ", "+")

    start = time.perf_counter()
    retries = 3

    for attempt in range(retries):
        try:
            params_query = query
            if not is_address:
                try:
                    lat, lon = [p.strip() for p in query.split(",", 1)]
                    params_query = f"{lon},{lat}"
                except Exception:
                    params_query = query

            response = requests.get(
                "https://geocode-maps.yandex.ru/v1/",
                params={
                    "apikey": os.getenv("YANDEX_API_KEY"),
                    "geocode": params_query,
                    "format": "json",
                    "lang": "ru_RU"
                },
                timeout=10,
            )

            elapsed = time.perf_counter() - start

            data = response.json()

            geo_object = data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]

            if is_address:
                result = " ".join(geo_object["Point"]["pos"].split()[::-1])
            else:
                postal = geo_object["metaDataProperty"]["GeocoderMetaData"].get(
                    "postal_code", ""
                )
                result = f"{postal}, {geo_object['metaDataProperty']['GeocoderMetaData']['text']}"

            quality = geo_object["metaDataProperty"]["GeocoderMetaData"]["precision"]

            return result, quality, response.status_code, elapsed

        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)
                continue
            elapsed = time.perf_counter() - start
            print(f"[Yandex] Network error for query '{query}': {e}")
            return "", f"network error: {e}", 0, elapsed

def request_dadata(query, is_address):
    DADATA_TOKEN = os.getenv("DADATA_API_KEY")
    DADATA_SECRET = os.getenv("DADATA_SECRET")

    start = time.perf_counter()
    retries = 1

    for attempt in range(retries):
        try:
            # ----------------------------
            # Адрес -> координаты
            # ----------------------------
            if is_address:
                url = "https://cleaner.dadata.ru/api/v1/clean/address"

                headers = {
                    "Authorization": f"Token {DADATA_TOKEN}",
                    "X-Secret": DADATA_SECRET,
                    "Content-Type": "application/json",
                    "Accept": "application/json"
                }

                time.sleep(random.uniform(0.3, 0.7))

                response = requests.post(
                    url,
                    headers=headers,
                    json=[query],
                    timeout=10
                )

                response.raise_for_status()

                data = response.json()

                if not data:
                    elapsed = time.perf_counter() - start
                    return "", "Адрес не найден", response.status_code, elapsed

                item = data[0]

                lat = item.get("geo_lat")
                lon = item.get("geo_lon")

                elapsed = time.perf_counter() - start

                if not lat or not lon:
                    return "", "Координаты отсутствуют", response.status_code, elapsed

                result = f"{lat} {lon}"
                quality = item.get("qc_geo", "_")

                return result, quality, response.status_code, elapsed

            # ----------------------------
            # Координаты -> адрес
            # ----------------------------
            else:
                lat, lon = map(float, query.split(","))

                url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/geolocate/address"

                headers = {
                    "Authorization": f"Token {DADATA_TOKEN}",
                    "Content-Type": "application/json",
                    "Accept": "application/json"
                }

                payload = {
                    "lat": lat,
                    "lon": lon
                }

                print(payload)

                time.sleep(random.uniform(0.3, 0.7))

                response = requests.post(
                    url,
                    headers=headers,
                    json=payload,
                    timeout=10
                )

                response.raise_for_status()

                data = response.json()

                suggestions = data.get("suggestions", [])

                elapsed = time.perf_counter() - start

                if not suggestions:
                    return "", "Адрес не найден", "_", elapsed

                suggestion = suggestions[0]

                result = suggestion.get(
                    "unrestricted_value",
                    suggestion.get("value", "")
                )

                quality = suggestion.get("data", {}).get("qc_geo", "_")

                print(quality)

                return result, quality, response.status_code, elapsed

        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)
                continue

            elapsed = time.perf_counter() - start

            print(f"[DaData] Network error for query '{query}': {e}")

            return "", f"network error: {e}", 0, elapsed


def request_gigachat(query, is_address):
    if is_address:
        prompt = f"""
        Определи географические координаты адреса:

        {query}

        Ответь ТОЛЬКО двумя числами в формате:
        latitude longitude

        Никаких пояснений, текста, единиц измерения, markdown и дополнительных символов.
        """
    else:
        prompt = f"""
        Определи полный почтовый адрес по координатам:

        {query}

        Ответь ТОЛЬКО адресом.
        Никаких пояснений, вводных фраз и дополнительного текста.
        """

    headers = {
        "Authorization": f"Bearer {os.getenv('GIGACHAT_TOKEN')}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    payload = {
        "model": "GigaChat-2",
        "messages": [
            {
                "role": "user",
                "content": prompt
            }
        ],
        "stream": False,
        "repetition_penalty": 1
    }

    start = time.perf_counter()
    retries = 1

    for attempt in range(retries):
        try:
            time.sleep(random.uniform(0.3, 0.7))

            response = requests.post(
                "https://api.giga.chat/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=10,
                verify=False
            )

            response.raise_for_status()

            elapsed = time.perf_counter() - start

            data = response.json()

            result = (
                data["choices"][0]["message"]["content"]
                .strip()
            )

            return result, "-", response.status_code, elapsed

        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)
                continue

            elapsed = time.perf_counter() - start

            print(f"[GigaChat] Network error for query '{query}': {e}")

            return "", f"network error: {e}", 0, elapsed

def calculate_statistics():

    stats = defaultdict(list)

    with open("results.csv", encoding="utf-8-sig") as file:

        reader = csv.DictReader(file)

        for row in reader:

            service = row.get("Service", "")
            operation = row.get("Operation", "")

            response_time = float(row["Response Time (ms)"])

            stats[(service, operation)].append(response_time)

    with open("statistics.csv", "w", newline="", encoding="utf-8-sig") as file:

        writer = csv.writer(file)

        writer.writerow([
            "Service",
            "Operation",
            "Requests",
            "Average (ms)",
            "Min (ms)",
            "Max (ms)"
        ])

        for (service, operation), values in stats.items():

            writer.writerow([
                service,
                operation,
                len(values),
                round(sum(values) / len(values), 2),
                round(min(values), 2),
                round(max(values), 2),
            ])


def generate_accuracy():
    results = defaultdict(lambda: {"total": 0, "success": 0})

    with open("results.csv", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)

        for row in reader:
            service = (row.get("Service") or "").strip().lower()
            operation = (row.get("Operation") or "").strip()

            key = (service.capitalize(), operation)

            results[key]["total"] += 1

            quality = (row.get("Quality") or "").strip()
            result_val = (row.get("Result") or "").strip()
            status = (row.get("HTTP Status") or "").strip()

            try:
                status_code = int(status)
            except ValueError:
                status_code = 0

            success = False

            # ----------------------------
            # Yandex
            # ----------------------------
            if service == "yandex":
                if operation == "Direct":
                    success = quality.lower() in ("exact", "number", "near")
                else:
                    success = quality.lower() in ("exact", "number", "street")

            # ----------------------------
            # DaData
            # ----------------------------
            elif service == "dadata":
                if operation == "Direct":
                    # qc_geo:
                    # 0 - точные координаты дома
                    # 1 - ближайший дом
                    # 2 - улица
                    # 3 - населенный пункт
                    # 4 - город
                    # 5 - координаты отсутствуют
                    try:
                        success = int(quality) <= 1
                    except ValueError:
                        success = False
                else:
                    success = result_val != ""

            # ----------------------------
            # GigaChat
            # ----------------------------
            elif service == "gigachat":
                success = status_code == 200

            # ----------------------------
            # Остальные сервисы
            # ----------------------------
            else:
                success = result_val != ""

            if success:
                results[key]["success"] += 1

    with open("accuracy.csv", "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)

        writer.writerow([
            "Service",
            "Operation",
            "Total Requests",
            "Successful",
            "Failed",
            "Success %"
        ])

        for (service, operation), vals in sorted(results.items()):
            total = vals["total"]
            success = vals["success"]
            failed = total - success
            percent = round(success / total * 100, 2) if total else 0

            writer.writerow([
                service,
                operation,
                total,
                success,
                failed,
                percent
            ])

if __name__ == "__main__":

    warnings.filterwarnings("ignore")

    load_dotenv()

    try:
        update_gigachat_token()
    except Exception as e:
        print(f"GigaChat token error: {e}")

    addresses, coordinates = read_excel("Геоданные.xlsx")

    print("Address geocoding...")

    for address in addresses:

        result = request_yandex(address, True)
        save_result("Yandex", "Direct", address, *result)

        result = request_dadata(address, True)
        save_result("DaData", "Direct", address, *result)

        result = request_gigachat(address, True)
        save_result("GigaChat", "Direct", address, *result)

        time.sleep(1)

    print("Reverse geocoding...")

    for coords in coordinates:

        result = request_yandex(coords, False)
        save_result("Yandex", "Reverse", coords, *result)

        result = request_dadata(coords, False)
        save_result("DaData", "Reverse", coords, *result)

        result = request_gigachat(coords, False)
        save_result("GigaChat", "Reverse", coords, *result)

        time.sleep(1)

    print("Done!")

    calculate_statistics()
    print("Statistics saved.")

    try:
        generate_accuracy()
        print("Accuracy saved.")
    except Exception as e:
        print(f"Accuracy generation error: {e}")